import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

def create_excel(data):
    try:
        book = openpyxl.load_workbook('output.xlsx')
    except FileNotFoundError:
        book = openpyxl.Workbook()
        book.remove(book.active)

    for task_id, task_info in data.items():
        filename = 'output.xlsx'
        column = task_info['column']
        title = task_info['title']
        description = task_info['description']
        stickers = task_info['stickers']

        sticker_info = {sticker.split(':')[0]: sticker.split(':')[1].strip() for sticker in stickers if ':' in sticker}

        if column == 'Общее':
            df = pd.DataFrame({'Название': [title], 'Описание': [description]})
            if 'Общее' in book.sheetnames:
                sheet = book['Общее']
            else:
                sheet = book.create_sheet('Общее')
                for row in dataframe_to_rows(df.iloc[0:0], index=False, header=True):
                    sheet.append(row)
        elif column == 'Действия':
            df = pd.DataFrame({'Название': [title], 'Описание': [description], 'Количество карт': [sticker_info.get('Количество карт')], 'Поле': [sticker_info.get('Поле')], 'Стоимость': [sticker_info.get('Стоимость')], 'Автор': [sticker_info.get('Автор')]})
            if 'Действия' in book.sheetnames:
                sheet = book['Действия']
            else:
                sheet = book.create_sheet('Действия')
                for row in dataframe_to_rows(df.iloc[0:0], index=False, header=True):
                    sheet.append(row)
        elif column in ['Пассивные эффекты', 'Защитные эффекты']:
            df = pd.DataFrame({'Название': [title], 'Описание': [description],'Количество карт': [sticker_info.get('Количество карт')], 'Автор': [sticker_info.get('Автор')]})
            if column in book.sheetnames:
                sheet = book[column]
            else:
                sheet = book.create_sheet(column)
                for row in dataframe_to_rows(df.iloc[0:0], index=False, header=True):
                    sheet.append(row)
        elif column in ['Персонажи', 'Остальные идеи']:
            df = pd.DataFrame({'Название': [title], 'Описание': [description], 'Автор': [sticker_info.get('Автор')]})
            if column in book.sheetnames:
                sheet = book[column]
            else:
                sheet = book.create_sheet(column)
                for row in dataframe_to_rows(df.iloc[0:0], index=False, header=True):
                    sheet.append(row)
        elif column == 'События':
            df = pd.DataFrame({'Название': [title], 'Описание': [description], 'Влияние': [sticker_info.get('Влияние')], 'Тип': [sticker_info.get('Тип')],'Количество карт': [sticker_info.get('Количество карт')], 'Автор': [sticker_info.get('Автор')]})
            if 'События' in book.sheetnames:
                sheet = book['События']
            else:
                sheet = book.create_sheet('События')
                for row in dataframe_to_rows(df.iloc[0:0], index=False, header=True):
                    sheet.append(row)

        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adj_width = (max_length + 5)
            sheet.column_dimensions[column].width = adj_width

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    book.save(filename)
    print(f'{filename} успешно создан! ;)')